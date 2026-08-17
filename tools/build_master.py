#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BNCP 대시보드 데이터 빌드
  입력: BNCP_공종마스터_컬럼통일.xlsx
  출력: assets/js/master.js   (공종마스터 975코드)
        assets/js/materials.js(자재원단위 44시트 → long table)
원단위가 바뀌면 이 스크립트만 다시 돌린다.
"""
import json, re, sys, os
import openpyxl

SRC = sys.argv[1] if len(sys.argv) > 1 else 'BNCP_공종마스터_컬럼통일.xlsx'
OUT = sys.argv[2] if len(sys.argv) > 2 else 'assets/js'

SKIP = {'변환검증', '삭제_비토목_GL'}

# 자재원단위가 아니라 '관경별 굴착 치수 참조표 / 계산기' — 자재 목록에서 뺀다
REF_SHEETS = {
    '관경별_규격표(우수관)', '관경별_규격표(오수관)', '관경별_규격표(오수관_보도)',
    '우수관_터파기계산기(아스팔트구간)', '오수관_터파기계산기(아스팔트+녹지)',
    '오수관_터파기계산기(보도구간)',
    '상수도공_규격표_상수관', '상수도공_규격표_상수관_보도', '상수도공_계산기_상수관',
    '관수공_규격표_관로', '관수공_규격표_관로_보도',
}
# 지급 대상 자재가 아닌 항목 — 거푸집(전용 가설재) · 토공 물량 · 치수값
NON_MAT = re.compile(
    r'거푸집|터파기|되메우기|잔토|표토제거|원지반|'
    r'외경|저폭|상단폭|토피고|포장두께|기초\s*Hb|골재치환|관길이|본당환산|'
    r'표식테이프|기초종류|관종|모래단위|연장|구간|'
    r'^관경$|보호공길이|항목명 미기재')

def clean_name(v):
    """'5. 거푸집' → '거푸집',  '-' → 빈값"""
    t = re.sub(r'^\s*\d+\s*[.)]\s*', '', s(v))
    return '' if t in ('-', '', 'ㅡ') else t
MASTER_SHEETS = ['공종마스터', '부대토목_공종마스터']

wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)


def s(v):
    if v is None:
        return ''
    t = str(v).strip()
    return re.sub(r'\s+', ' ', t)


def num(v):
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return round(float(v), 6)
    t = str(v).replace(',', '').strip()
    try:
        return round(float(t), 6)
    except ValueError:
        return None


# ────────────────────────────── 1. 공종마스터 ──────────────────────────────
codes, dup = [], set()
for sh in MASTER_SHEETS:
    ws = wb[sh]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [s(c) for c in rows[0]]
    ix = {h: i for i, h in enumerate(hdr)}
    for r in rows[1:]:
        code = s(r[ix['코드']])
        if not code:
            continue
        if code in dup:
            print('  ! 중복코드', code)
            continue
        dup.add(code)
        codes.append({
            'c': code,
            'g': s(r[ix['대분류']]),
            'm': s(r[ix['중분류']]),
            'n': s(r[ix['공종명(leaf)']]),
            'sp': s(r[ix['규격']]),
            'u': s(r[ix['단위']]),
            'src': s(r[ix['출처블록']]),
        })
print(f'공종마스터 {len(codes)}코드')

# 대분류 순서(엑셀 등장 순)
order, seen = [], set()
for c in codes:
    if c['g'] not in seen:
        seen.add(c['g'])
        order.append(c['g'])


# ────────────────────────────── 2. 자재원단위 ──────────────────────────────
def find_header(rows):
    """3개 이상 값이 있는 첫 행을 헤더로 본다."""
    for i, r in enumerate(rows[:4]):
        if sum(1 for c in r if s(c)) >= 3:
            return i, [s(c) for c in r]
    return 0, [s(c) for c in rows[0]]


def split_unit(label):
    """'레미콘\\n철근31\\n(㎥)' → ('레미콘 철근31', '㎥')"""
    m = re.search(r'\(([^()]{1,12})\)\s*$', label)
    if m:
        return label[:m.start()].strip(' ·-'), m.group(1)
    return label, ''


LONG_KEYS = {'재료', '공종', '세부공종', '세부항목'}
QTY_PAT = re.compile(r'(수량|원단위|단위수량)')

mats = []          # 지급 대상 자재
excl = []          # 자재 아님(거푸집·토공물량·치수값) — 근거 남기려고 보관
tables = []        # 규격표·계산기 원본 보존

for sh in wb.sheetnames:
    if sh in SKIP or sh in MASTER_SHEETS:
        continue
    ws = wb[sh]
    rows = [r for r in ws.iter_rows(values_only=True)]
    if not rows:
        continue
    hi, hdr = find_header(rows)
    if sh in REF_SHEETS:                    # 치수 참조표는 표 그대로 보존
        tables.append({'sh': sh, 'hdr': hdr,
                       'rows': [[s(c) for c in r] for r in rows[hi + 1:] if any(s(c) for c in r)]})
        continue
    body = rows[hi + 1:]
    qty_cols = [i for i, h in enumerate(hdr) if QTY_PAT.search(h)]
    has_long = any(h in LONG_KEYS for h in hdr)

    if has_long and qty_cols:
        # ── 세로형 ──
        ix = {h: i for i, h in enumerate(hdr)}
        qi = qty_cols[0]
        ui = ix.get('단위')
        for r in body:
            v = num(r[qi]) if qi < len(r) else None
            if v is None:
                continue
            grp = s(r[ix['구분']]) if '구분' in ix else s(r[0])
            item = ''
            for k in ('세부공종', '세부항목', '공종', 'Type'):
                if k in ix and s(r[ix[k]]):
                    item = s(r[ix[k]]); break
            mat = s(r[ix['재료']]) if '재료' in ix else item
            unit = s(r[ui]) if ui is not None and ui < len(r) else ''
            basis = s(r[ix['산출근거']]) if '산출근거' in ix else ''
            _, hu = split_unit(hdr[qi])
            name = clean_name(mat) or clean_name(item) or clean_name(grp)
            rec = {'sh': sh, 'k': clean_name(grp), 'item': clean_name(item), 'mat': name,
                   'u': unit, 'v': v, 'b': basis, 'per': hu or s(hdr[qi])}
            (excl if NON_MAT.search(name) else mats).append(rec)
    else:
        # ── 가로형(매트릭스) : 1열 = 규격키, 나머지 = 자재 ──
        key_label = hdr[0] or '규격'
        wide_ok = False
        for r in body:
            key = s(r[0])
            if not key:
                continue
            for ci in range(1, len(hdr)):
                lab = hdr[ci]
                if not lab or ci >= len(r):
                    continue
                v = num(r[ci])
                if v is None or v == 0:
                    continue
                mat, unit = split_unit(lab)
                mat = clean_name(mat)
                rec = {'sh': sh, 'k': key_label, 'item': clean_name(key),
                       'mat': mat, 'u': unit, 'v': v, 'b': '',
                       'per': '개소/구간'}
                (excl if NON_MAT.search(mat) else mats).append(rec)
                wide_ok = True
        if not wide_ok:
            # 값이 안 잡히면 표 원본을 보존
            tables.append({'sh': sh, 'hdr': hdr,
                           'rows': [[s(c) for c in r] for r in body if any(s(c) for c in r)]})

print(f'자재 {len(mats)}행 / 제외 {len(excl)}행 / 참조표 {len(tables)}개')
ex = {}
for m in excl:
    ex[m['mat']] = ex.get(m['mat'], 0) + 1
print('  [자재 아님으로 제외한 항목]')
for k, v in sorted(ex.items(), key=lambda x: -x[1]):
    print(f'    {v:>4}  {k}')
by = {}
for m in mats:
    by[m['sh']] = by.get(m['sh'], 0) + 1
for k, v in sorted(by.items(), key=lambda x: -x[1]):
    print(f'   {k:36} {v}')

# ────────────────────────────── 3. 출력 ──────────────────────────────
os.makedirs(OUT, exist_ok=True)
with open(f'{OUT}/master.js', 'w', encoding='utf-8') as f:
    f.write('/* 자동생성 — build_master.py. 직접 고치지 말 것 */\n')
    f.write('window.BNCP_MASTER=' + json.dumps(
        {'codes': codes, 'order': order}, ensure_ascii=False, separators=(',', ':')) + ';\n')

with open(f'{OUT}/materials.js', 'w', encoding='utf-8') as f:
    f.write('/* 자동생성 — build_master.py. 직접 고치지 말 것 */\n')
    f.write('window.BNCP_MAT=' + json.dumps(
        {'rows': mats, 'tables': tables}, ensure_ascii=False, separators=(',', ':')) + ';\n')

print('완료 →', OUT)
