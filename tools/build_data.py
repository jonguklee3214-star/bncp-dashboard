# -*- coding: utf-8 -*-
"""data.js 재생성 — 업로드된 xlsx + SKILL.md 원본에서 전부 다시 뽑는다.
(이전 세션의 data.js가 첨부되지 않아 원본에서 복원)"""
import openpyxl, json, re, os
from collections import OrderedDict

U = '/mnt/user-data/uploads'
def wbk(f): return openpyxl.load_workbook(os.path.join(U, f), data_only=True)
def num(x):
    if x is None: return 0.0
    if isinstance(x, (int, float)): return float(x)
    s = str(x).replace(',', '').strip()
    try: return float(s)
    except: return 0.0

D = OrderedDict()

# ---------------------------------------------------------------- 1. 생산성 PROD
PROD, PMETA = {}, {}
row_re = re.compile(r'^\|\s*([A-Z][A-Za-z0-9_가-힣]*)\s*\|(.+)\|\s*$')
for f in ['bncp-structure-schedule_SKILL.md', 'bncp-civil-schedule_SKILL.md',
          'bncp-ancillary-civil-schedule_SKILL.md']:
    src = f.split('_')[0]
    for line in open(os.path.join(U, f), encoding='utf-8'):
        m = row_re.match(line.rstrip())
        if not m: continue
        key = m.group(1)
        if key in ('Key', 'ID', 'Code'): continue
        cells = [c.strip() for c in m.group(2).split('|')]
        val, unit, name = None, '', cells[0] if cells else ''
        for c in cells[1:]:
            cc = c.replace('**', '').replace(',', '').strip()
            if re.fullmatch(r'\d+(\.\d+)?', cc):
                val = float(cc)
        for c in cells:
            if c in ('m³', 'm²', 'm', 'ea', 'ton', '개소', 'nr·ea', 'm²/m', '공/m³', '㎥', '㎡'):
                unit = c
        if val is None:
            PMETA[key] = {'n': name, 'u': unit, 's': src, 'warn': 1}
            continue
        if key not in PROD:          # 첫 정의 우선 (스킬 내 중복키 존재)
            PROD[key] = val
            PMETA[key] = {'n': name, 'u': unit, 's': src}
# 매핑표 표기 ↔ 스킬 Key_ID 표기 차이 보정
if 'ROW_012' in PROD:
    PROD['ROW_012(civil)'] = PROD['ROW_012']
    PMETA['ROW_012(civil)'] = dict(PMETA['ROW_012'], alias='ROW_012')
D['PROD'] = PROD
D['PMETA'] = PMETA

# ---------------------------------------------------------------- 2. 부지토목 공종 트리
wb = wbk('BNCP_공종마스터_컬럼통일.xlsx')
ws = wb['공종마스터']
CIV = OrderedDict()
for r in ws.iter_rows(min_row=2, values_only=True):
    code, bigc, big, mid, small, detail, leaf, spec, unit, note, blk = r[:11]
    if blk != '부지토목': continue
    CIV.setdefault(big, OrderedDict()).setdefault(mid or '-', []).append({
        'c': code, 'n': leaf or small or mid, 'sp': spec or '', 'u': unit or '', 'nt': note or ''})
D['CIVIL'] = CIV
D['CIVIL_ORDER'] = list(CIV.keys())

# 부대토목(참고용 · 블록단위)
ws2 = wb['부대토목_공종마스터']
ANC = OrderedDict()
for r in ws2.iter_rows(min_row=2, values_only=True):
    code, bigc, big, mid, small, detail, leaf, spec, unit, note, blk = r[:11]
    ANC.setdefault(big, OrderedDict()).setdefault(mid or '-', []).append({
        'c': code, 'n': leaf or small or mid, 'sp': spec or '', 'u': unit or ''})
D['ANC'] = ANC
D['ANC_ORDER'] = list(ANC.keys())

# ---------------------------------------------------------------- 3. 구조물 원단위
# GL — 원단위_매트릭스
wb = wbk('BNCP_갤러리_원단위_계산기.xlsx')
ws = wb['원단위_매트릭스']
hdr = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
cols = [h for h in hdr[4:] if h]
rows = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r[0]: continue
    v = [num(x) for x in r[4:4+len(cols)]]
    if not any(v): continue
    rows.append({'g': r[0], 's': r[1] or '', 'u': r[3] or '', 'v': v})
D['GL'] = {'cols': cols, 'rows': rows}

# PS / SP / IT — 원단위_구조물공
for tag, fn, ncol in [('PS', 'BNCP_빗물펌프장_원단위.xlsx', 3),
                      ('SP', 'BNCP_오수중계펌프장_원단위.xlsx', 3),
                      ('IT', 'BNCP_이리게이션탱크_원단위.xlsx', 2)]:
    ws = wbk(fn)['원단위_구조물공']
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    # IT는 세부 컬럼이 없어 단위열 위치가 다름
    ui = 3 if tag != 'IT' else 2
    st = ui + 1
    cols = [h for h in hdr[st:st+ncol]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[0]: continue
        v = [num(x) for x in r[st:st+ncol]]
        if not any(v): continue
        rows.append({'g': r[0], 's': r[1] or '', 'u': r[ui] or '', 'v': v})
    D[tag] = {'cols': cols, 'rows': rows}

# ---------------------------------------------------------------- 4. 매핑표
ws = wbk('BNCP_원단위_생산성_매핑표.xlsx')['매핑표']
FACK = {'GL 갤러리': 'GL', '빗물펌프장': 'PS', '오수중계P/S': 'SP',
        '오수중계펌프장': 'SP', '이리게이션탱크': 'IT'}
MAP = {}
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r[0]: continue
    fac = None
    for k, v in FACK.items():
        if k in str(r[0]): fac = v; break
    if not fac: continue
    MAP[f'{fac}|{r[1]}|{r[2] or ""}'] = {'c': (r[4] or 'B').strip(),
                                         'k': (r[5] or '').strip(), 'n': (r[6] or '').strip()}
D['MAP'] = MAP

# ---------------------------------------------------------------- 5. 자원표
ws = wbk('BNCP_현장_실배치_자원표.xlsx')['실배치_자원표']
hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
eqcols = hdr[7:-1]
R = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r[0] or not r[1]: continue
    eq = {eqcols[i]: int(num(r[7+i])) for i in range(len(eqcols)) if num(r[7+i])}
    R.append({'site': r[0], 'g': r[1], 'w': r[2] or '', 'loc': r[3] or '',
              't': int(num(r[4])), 'pp': int(num(r[5])), 'p': int(num(r[6])),
              'eq': eq, 'et': int(num(r[-1]))})
D['RES'] = {'rows': R, 'eqcols': eqcols}
D['META'] = {'total': {'teams': sum(x['t'] for x in R), 'people': sum(x['p'] for x in R),
                       'equip': sum(x['et'] for x in R)},
             'team': next(({'people': x['p'], 'teams': x['t']} for x in R if '구조물' in x['g']),
                          {'people': 0, 'teams': 0})}

# ---------------------------------------------------------------- 6. 페이즈
# 기준 통일(§2-0-B) 결과 — 팀당 생산성
PB = json.load(open('/home/claude/bncp/skill/prod_basis.json', encoding='utf-8'))
if 'ROW_012' in PROD:
    PB['ROW_012(civil)'] = {'n': '잔토처리', 'u': 'm³', 'b': '조', 'per': None,
                            'N': 1, 'team': PROD['ROW_012']}
D['PB'] = PB

D['PHASES'] = ['Phase 1', 'Phase 2', 'Phase 3-1', 'Phase 3-2', 'Phase 4']

out = 'window.BNCP=' + json.dumps(D, ensure_ascii=False, separators=(',', ':')) + ';\n'
open('/home/claude/bncp/data.js', 'w', encoding='utf-8').write(out)

print('PROD keys :', len(PROD))
print('CIVIL 대분류:', list(CIV.keys()))
print('CIVIL leaves:', sum(len(v) for m in CIV.values() for v in m.values()))
for t in ['GL', 'PS', 'SP', 'IT']:
    print(t, 'cols=', D[t]['cols'], 'rows=', len(D[t]['rows']))
print('MAP :', len(MAP), ' RES rows:', len(R), D['META'])
print('bytes:', len(out))
